# @author: Z
# @time: 2021/7/22 12:22
# @desc:Excel读写
import os
import openpyxl

from Rconfig.configRW import *

# def Getconfigdata():
#     a = configRW().configR('Excel')
#     return a

class RWexcel:

    def __init__(self, sheet):
        # 读取配置信息
        a = configRW().configR('Excel')
        self.workbook=a['workbook']
        self.sheet=sheet
        # print(a)
        # for b in a:
        #     if b.get('workbook'):
        #         self.workbook=b.get('workbook')
        #     if b.get('sheet1'):
        #         self.sheet = b.get('sheet1')


    # # Excel读数据
    def Rexcel(self):
            # RWexcel.sheet=sheets
            # print(RWexcel.sheet)
            workbook = openpyxl.load_workbook(self.workbook)
            sheet = workbook[self.sheet]
            r_max = sheet.max_row
            # print(r_max)
            list_ = []
            for i in range(2, r_max + 1):

                data = dict(id=sheet.cell(row=i, column=1).value,
                            url=sheet.cell(row=i, column=2).value,
                            headers=sheet.cell(row=i, column=3).value,
                            data=sheet.cell(row=i, column=4).value,
                            methon=sheet.cell(row=i, column=5).value,
                            excett=sheet.cell(row=i, column=6).value,
                            用例名=sheet.cell(row=i, column=8).value)

                list_.append(data)
            return list_

    # Excel写数据
    def Wexcel(self, r, c, data):
        # print('----------------'+self.sheet)
        workbook = openpyxl.load_workbook(self.workbook)
        sheet = workbook[self.sheet]
        sheet.cell(r, c).value=data
        workbook.save(self.workbook)

